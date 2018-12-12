import time
#导入操作excel文件的函数，如果没有安装，可通过pip install openpyxl安装
from openpyxl import workbook
from openpyxl import load_workbook
import os.path
from common.loggen import Logger
from selenium import webdriver
from pages.mainpage import MainPage
from pages.basepage import BasePage
from pages.loginpage import LoginPage
from pages.registepage import RegistePage
from common.geturl import geturl
logger = Logger(logger="TestSuite").getlog()
#创建读取测试集函数
def read_testsuite(tsname):
    #设置测试用例读取执行状态标志位
    flag = True
    #设置读取测试集函数执行状态标志位
    read_testsuite = True
    #判断测试集文件是否存在
    if os.path.exists(tsname):
        #如果存在则写入日志
        logger.info('已找到TestSuite文件，开始分析测试集...')
        #创建excel操作对象
        wbexcel = load_workbook(tsname)
        sheetnames = wbexcel.get_sheet_names()
        ws = wbexcel.get_sheet_by_name(sheetnames[0])
        #分析测试集文件中的执行信息：执行标志位及测试脚本名称，从第二行开始
        for irow in range(2,ws.max_row+1):
            #获取测试集文件中的执行标志位值，从第二行开始，第二列
            testoperation = ws.cell(row=irow, column=2).value
            #获取测试集文件中的测试用例名称，从第二行开始，第三列
            testcasefile = ws.cell(row=irow, column=3).value
            #判断执行标志位是否需要执行，如果是do，则调用测试用例执行函数，如果是not，则不执行，如果是其他
            #值，则写入日志，报告执行参数错误，并指出是哪个用例执行参数错误
            if testoperation == 'do':
                logger.info('******************************')
                logger.info('执行 %s 测试场景' %testcasefile )
                #加载测试用例读取函数，并返回其返回值，以判断用例读取情况
                flag=read_testcase(testcasefile)
                #如果用例读取函数返回为False，则说明用例读取错误
                if flag==False:
                    logger.info('测试用例执行失败')
            #如果执行状态为not，说明当前用例无须执行
            elif testoperation == 'not':
                logger.info('%s 场景无须测试' % testcasefile)
            #如果既不是do，又不是not，则报告错误
            else:
                logger.info('执行参数错误，请检查%s' %testcasefile)
            #如果执行状态错误，则跳出循环，停止测试
                break
    #如果测试集文件错误，则写入日志，并提示错误原因
    else:
        logger.info('未发现:%s，请检查文件是否正确' % tsname)
        #返回测试集执行函数状态，便于run.py中的unittest中记录该状态
        read_testsuite = False
    #返回测试集执行函数执行状态
    return read_testsuite
#定义浏览器启动函数，本次并没有使用common中定义的browserlauncher函数，读者可自行扩展改写
def get_driver(testpage,teststep,testdata):
    #设置浏览器启动函数执行状态，便于后续运行控制
    get_driver = True
    #判断测试用例中是否需要启动浏览器，如果需要，则判断启动哪种浏览器
    if testpage == '浏览器':
        #考虑测试用例中的step大小写问题，读者自行研究解决
        if teststep == 'firefox':
            driver = webdriver.Firefox()
        elif teststep == 'ie':
            driver = webdriver.Ie()
        elif teststep == 'chrome':
            driver = webdriver.Chrome()
        #如果浏览器类型设置错误，写入日志并给予提示
        else:
            logger.info('未知浏览器类型，请检查测试用例')
        #启动没有问题后加载测试路径并返回driver对象
        driver.get(testdata)
        get_driver = driver
    else:
        #如果测试用例中的启动参数错误，则写入日志并给予提示
        logger.info('浏览器数据错误，请检查测试用例配置')
        get_driver = False
    return get_driver
#定义测试用例执行函数，共有四个参数
def exec_script(driver,testpage, teststep, testdata):
    #定义测试用例执行函数状态标志位
    exec_script = True
    try:
        #登陆功能测试
        if testpage == '登录':
            url = driver.current_url
            url = geturl(url) + 'user.php'
            if driver.current_url != url:
                driver.get(url)
            login = LoginPage(driver, testdata)
            if teststep == '用户名':
                login.input_username(testdata)

            if teststep == '密码':
                login.input_password(testdata)

            if teststep == '登录':
                login.click_submit()
                time.sleep(5)

        #注册功能测试
        if testpage == '注册':
            url = driver.current_url
            url = geturl(url) + 'user.php?act=register'
            if driver.current_url != url:
                driver.get(url)
            userreg = RegistePage(driver, testdata)
            if teststep == '用户名':
                userreg.input_username(testdata)

            if teststep == 'email':
                userreg.input_email(testdata)

            if teststep == '密码':
                userreg.input_password(testdata)

            if teststep == '确认密码':
                userreg.input_comfirpwd(testdata)

                time.sleep(8)
            if teststep == '注册':
                userreg.click_submit()
                time.sleep(5)

        if testpage == '主页':
            time.sleep(3)
            url = driver.current_url
            mainpage = MainPage(driver, url)
            if teststep == '退出':
                mainpage.exit_sys()
                time.sleep(3)
        if testpage == '其他主页':
            pass
    except:
        exec_script = False
        url = geturl(driver.current_url)
        driver.get(url)
    return exec_script

#定义测试用例读取函数
def read_testcase(testcasefile):
    #设置测试用例读取函数状态标志位
    read_testcase = True
    #根据read_testsuite函数中给出的testcasefile测试用例名，拼接测试用例路径信息
    testcasefile=os.path.abspath('.')+'\\data\\'+testcasefile+'.xlsx'
    #判断需读取执行的测试用例文件是否存在
    if os.path.exists(testcasefile):
        #如果存在，则写日志，并读取该用例的excel文件
        logger.info('已找到 %s 测试用例，现在开始读取该用例' %testcasefile)
        wbexcel = load_workbook(testcasefile)
        sheetnames = wbexcel.get_sheet_names()
        ws = wbexcel.get_sheet_by_name(sheetnames[0])
        #读取测试用例中每个列的值，以便调用浏览器启动函数或执行测试用例函数
        for irow in range(2, ws.max_row + 1):
            testpage = ws.cell(row=irow, column=1).value
            teststep = ws.cell(row=irow, column=2).value
            testdata = ws.cell(row=irow, column=4).value
            #如果是浏览器，说明需启动浏览器，调用浏览器启动函数
            if testpage=='浏览器':
                logger.info('正在启动浏览器')
                testdriver=get_driver(testpage, teststep, testdata)
            else:
            #如果不是浏览器，则说明需执行测试用例，调用测试用例执行函数
                flag=exec_script(testdriver,testpage, teststep, testdata)
        #执行完成后退出浏览器
        testdriver.quit()
    else:
        #如果测试用例文件不存在，则写入日志，并提示检查文件是否存在
        logger.info('未发现 %s 测试用例，请确认该用例是否存在' %testcasefile)
        #测试用例读取失败，状态标志位设置为False
        read_testcase = False
    #返回测试用例读取函数的状态，便于read_testsuite函数调用判断
    return read_testcase